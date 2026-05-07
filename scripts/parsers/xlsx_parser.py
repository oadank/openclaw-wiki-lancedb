"""parsers/xlsx_parser.py — Excel 表格解析（openpyxl）"""

def parse_xlsx(filepath):
    """将 .xlsx 解析为 Markdown 表格"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f'## 工作表: {sheet_name}\n')
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 第一行作表头
        headers = [str(c) if c is not None else '' for c in rows[0]]
        parts.append('| ' + ' | '.join(headers) + ' |')
        parts.append('| ' + ' | '.join(['---'] * len(headers)) + ' |')
        for row in rows[1:]:
            cells = [str(c) if c is not None else '' for c in row]
            parts.append('| ' + ' | '.join(cells) + ' |')
    wb.close()
    return '\n\n'.join(parts)
